import subprocess
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import time
from DataFileControl import DataFileControl
from tkinter import *
from openpyxl.styles.borders import Border, Side

import matplotlib
matplotlib.use('TkAgg')

#This class handles file IO 

class FileHandler():
    
    #constructor checks if used in system class or different class 
    def __init__(self,usedIn):
        if(str(type(usedIn))!="<class 'System.System'>"):
            self.system = usedIn.system
        else:
            self.system = usedIn

    #creates a timestamped datafile in spreadsheet format 
    def createNewDataFile(self):
        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y:%H:%M:%S")
        fname= "blood_analysis_"+dt_string+".xlsx"
        path = self.system.directory+fname
        self.system.dataFilePaths.append(path)#adds file path to dataFilePaths variable
        
        #creares a new workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "bloodcount"
        self.writeBoilerPlate(ws1)#writes headers and slide numbers
        
        #resizes cells to fit text, centers text, and zooms for better touch screen display 
        self.reSizeCells(ws1)
        self.centerText(ws1)
        ws1.sheet_view.zoomScale = 225
        wb.save(path)#saves datafile
        self.system.currFileIndex += 1#increments current file index variable
        
    #opens current datafile using currFileIndex variable     
    def openCurrentDataFile(self):
        
        #only open if file exists
        if(len(self.system.dataFilePaths)>0):
            subprocess.Popen(['xdg-open', self.system.dataFilePaths[self.system.currFileIndex]])
            
            time.sleep(2)#delay for file controls and keyboard to populate as well
            
            #create and open datafile controls
            r = Tk()
            r.geometry("400x150+1520+0")
            r.call('wm', 'attributes', '.', '-topmost', '1') #keeps the keypad on top
            r.title("Save/Close")
            r.overrideredirect(True)
            dfc = DataFileControl(r,self.system)
            dfc.grid()
            
    #writes headers and slide nubmers 
    def writeBoilerPlate(self,ws1):
        header = [u'Slide Position', u'Sample ID', u'Sample Date',
                    u'Analysis Date',u'Analysis Time',
                    u'WBC/RBC Ratio', u'Pathology']
        ws1.append(header)
        for i in range (20):
            ws1.cell(row=i+2, column=1, value=i)
            
    #writes ratio from analysis and adds red border if pathology is out of range 
    def writeRatio(self,wb,ws1,index,ratio,pathFlag):
        thin_border = Border(left=Side(style='thick',color='00FF0000'), 
                     right=Side(style='thick',color='00FF0000'), 
                     top=Side(style='thick',color='00FF0000'), 
                     bottom=Side(style='thick',color='00FF0000'))
        ws1.cell(row=index+2, column=6, value= ratio)
        if(pathFlag == True):
            ws1.cell(row=index+2, column=6).border = thin_border
        wb.save(self.system.dataFilePaths[self.system.currFileIndex])
#         print("~~~~~~~~~~~~~~~~~~Ratio written~~~~~~~~~~~~~~~~")


    #reads sample ID for GUI display for current slide position if sample ID is present in the datafile 
    def readSampleID(self):
        try:
            wb = load_workbook(self.system.dataFilePaths[self.system.currFileIndex])
            ws1 = wb.active
            sampleID = ws1.cell(row=self.system.carousel.curPos+2 ,column = 2).value
            print("SampleID is",sampleID)
            print("current Carousel Position is: ",self.system.carousel.curPos )
            print(type(sampleID))
            return sampleID
        except:
            print("Error could not read sampleID")
            return None
        
    #reads sample date for GUI display for current slide position if sample ID is present in the datafile 
    def readSampleDate(self):
        #assuming only two date formats entered: mm/dd/yyyy or mm-dd-yyyy
        try:
            wb = load_workbook(self.system.dataFilePaths[self.system.currFileIndex])
            ws1 = wb.active
            temp = ws1.cell(row=self.system.carousel.curPos+2 ,column = 3).value
            if(':' in str(temp)):
                temp = str(temp).split(' ')
                temp = temp[0]
                temp = temp.split('-')
                if(temp[1][0]=='0'):
                    temp[1] = temp[1][1:]
                if(temp[2][0]=='0'):
                    temp[2] = temp[2][1:]
                temp[0] = temp[0][2:]
                sampleDate="{}/{}/{}".format(temp[1],temp[2],temp[0])
            else:
                temp = str(temp).split('-')
                sampleDate="{}/{}/{}".format(temp[0],temp[1],temp[2])
                
            return sampleDate
        except:
            print("Error could not read sampleDate")
            return None
    
    #writes data and time of analysis to datafile 
    def writeDateTime(self,wb,ws1,index):
        now_date = datetime.now()
        temp = now_date.strftime("%d/%m/%Y")
        temp= temp.split('/')
        if(temp[0][0]=='0'):
            temp[0] = temp[0][1:]
        if(temp[1][0]=='0'):
            temp[1] = temp[1][1:]
        temp[2] = temp[2][2:]
        currDate="{}/{}/{}".format(temp[0],temp[1],temp[2])
        
        now_time = str(now_date)
        temp= now_time.split(' ')
        temp = str(temp[1])
        currTime = str(temp.split('.')[0])
#         print("~~~~~~~~~~~~~~~~~~{} {}~~~~~~~~~~~~~~~~".format(currDate,currTime))
        ws1.cell(row=index+2, column=4, value= currDate)#date
        ws1.cell(row=index+2, column=5, value= currTime)#time
        wb.save(self.system.dataFilePaths[self.system.currFileIndex])
#         print("~~~~~~~~~~~~~~~~~~DateTime written~~~~~~~~~~~~~~~~")
        
    #Writes pathology based on ratio and pathology min and max 
    def writePathology(self,wb,ws1,index,pathFlag):
        if(pathFlag==True):
            ws1.cell(row=index+2, column=7, value= "Out of Range")
            
        else:
            ws1.cell(row=index+2, column=7, value= "Normal")
            
        wb.save(self.system.dataFilePaths[self.system.currFileIndex])
        
        
    #resizes columns to fit headers    
    def reSizeCells(self, ws1):
        ws1.column_dimensions['A'].width= 13
        ws1.column_dimensions['B'].width= 10
        ws1.column_dimensions['C'].width= 12
        ws1.column_dimensions['D'].width= 14
        ws1.column_dimensions['E'].width= 14
        ws1.column_dimensions['F'].width= 14
        ws1.column_dimensions['G'].width= 12

    #centers text 
    def centerText(self,ws1):
        for col in ws1.columns:
            for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj