from FileHandler import FileHandler
from openpyxl import load_workbook

#This class handles functionality from the manual frame in the GUI
#does blood analysis, and saves data
#also has a Filehandler object

class Manual():
    
    #constructor is passed the system object
    def __init__(self,system):
        self.system = system
        self.fileHandler = FileHandler(self)
        self.bloodData = []
    
    #Is called is a separate thread. Does blood analysis and returns data as an array, also sends data to mainloop(GUI thread) via the shared queue
    def countBlood(self):
        #switched order to RBC counted first
        wbc = self.system.bloodCounter.countWBC(self.system.currImage_analyze)
        rbc = self.system.bloodCounter.countRBC(self.system.currImage_analyze)
        ratio = self.system.bloodCounter.calcRatio()
            
        self.bloodData = [wbc,rbc,ratio]
        pathFlag = self.system.util.pathologyWarn(self.bloodData[2])
        dataArray = ['countblood',self.bloodData[0], self.bloodData[1],self.bloodData[2],pathFlag,self.system.carousel.curPos]
        self.system.GUI.queue.put(dataArray)
        return self.bloodData

    #saves data to the data file,creates the file if it doesn't exist 
    def saveData(self):
        if(len(self.system.dataFilePaths)==0 and len(self.bloodData) == 3): #if file no exist and blood data exist, create new file
            self.system.util.createFile()
        
        try:
            pathFlag = self.system.util.pathologyWarn(self.bloodData[2])
            self.wb = load_workbook(self.system.dataFilePaths[self.system.currFileIndex])
            self.ws1 = self.wb.active
            self.fileHandler.writeRatio(self.wb,self.ws1,(self.system.carousel.curPos),self.bloodData[2],pathFlag)
            self.fileHandler.writeDateTime(self.wb,self.ws1,(self.system.carousel.curPos))
            self.fileHandler.writePathology(self.wb,self.ws1,(self.system.carousel.curPos),pathFlag)
        except:
            print("Could not load_workbook or write to file")

  
        
        
        
        
    
        