import subprocess
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os.path 
from datetime import datetime

class Utility:
    def __init__(self):
        self.minPathology = 0
        self.maxPathology = 0
        self.directory ="/home/pi/BAS/DataFiles/"
        self.paths =[]
        self.cnt = -1
    def setPathology(self, min, max):
        self.minPathology = min
        self.maxPathology = max
    def createNewDataFile(self):
        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y:%H:%M:%S")
        fname= "bloodcount_"+dt_string+".xlsx"
        path = self.directory+fname
        self.paths.append(path)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "bloodcount"
        self.writeBoilerPlate(ws1)
        self.reSizeCells(ws1)
        self.centerText(ws1)
        wb.save(path)
        subprocess.Popen(['xdg-open',path])
        self.cnt = self.cnt + 1
    def openCurrentDataFile(self):
        if(len(self.paths)>0):
            subprocess.Popen(['xdg-open',self.paths[self.cnt]])
    def writeBoilerPlate(self,ws1):
        header = [u'Slide Position', u'Sample ID', u'Sample Date',
                    u'Analysis Date',u'Analysis Time',
                    u'WBC/RBC Ratio', u'Pathology']
        ws1.append(header)
        for i in range (1,21):
            ws1.cell(row=i+1, column=1, value=i)
    def reSizeCells(self, ws1):
        column_widths = []
        for row in ws1.rows:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))
                else:
                    column_widths += [len(str(cell.value))]

        for i, column_width in enumerate(column_widths):
            ws1.column_dimensions[get_column_letter(i+1)].width= column_width+1
    def centerText(self,ws1):
         for col in ws1.columns:
             for cell in col:
            # openpyxl styles aren't mutable,
            # so you have to create a copy of the style, modify the copy, then set it back
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj
            
        
        
    
        
        

    
    